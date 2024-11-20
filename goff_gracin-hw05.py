# Gracin Goff
# UWYO COSC 1010
# 11/19/2024
# HW 05
# Lab Section: 10
# Sources, people worked with, help given to: 
# your
# comments
# here


import openpyxl 
from openpyxl.styles import Color , PatternFill
import string 
wb = openpyxl.Workbook()
sheet = wb.active 
background = Color(rgb='ADD8E6')
fill = PatternFill(patternType = 'solid', fgColor = background)
body = Color(rgb='00008B')
fill_b = PatternFill(patternType = 'solid', fgColor = body)
eyes = Color(rgb='000000')
fill_e = PatternFill(patternType = 'solid', fgColor = eyes)
white = Color(rgb='FFFFFF')
fill_w = PatternFill(patternType = 'solid', fgColor = white)
mouth = Color(rgb = 'FF0000')
fill_r = PatternFill(patternType = 'solid', fgColor = mouth)

cells_lb = ['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1',
            'A2','B2','C2','D2','E2','F2','G2','H2','I2','J2','K2','L2','M2','N2','O2','P2','Q2','R2','S2','T2','U2',
            'A3','B3','C3','D3','E3','F3','G3','H3','I3','J3','K3','L3','M3','N3','O3','P3','Q3','R3','S3','T3','U3',
            'A4','B4','C4','D4','E4','F4','G4','H4','I4','J4','L4','M4','N4','O4','P4','Q4','R4','S4','T4','U4',
            'A5','B5','C5','D5','E5','F5','G5','H5','I5','M5','N5','O5','P5','Q5','R5','S5','T5','U5',
            'A6','B6','C6','D6','E6','F6','G6','O6','P6','Q6','R6','S6','T6','U6',
            'A7','B7','C7','D7','E7','F7','P7','Q7','R7','S7','T7','U7',
            'A8','B8','C8','D8','E8','Q8','R8','S8','T8','U8',
            'A9','B9','C9','D9','R9','S9','T9','U9',
            'A10','B10','C10','D10','R10','S10','T10','U10',
            'A11','B11','C11','S11','T11','U11',
            'A12','B12','C12','S12','T12','U12',
            'A13','B13','T13','U13',
            'A14','B14','T14','U14',
            'A15','B15','T15','U15',
            'A16','B16','T16','U16',
            'A17','B17','T17','U17',
            'A18','B18','C18','D18','E18','F18','G18','H18','I18','J18','K18','L18','M18','N18','O18','P18','Q18','R18','S18','T18','U18',
            'A19','B19','C19','D19','E19','F19','G19','H19','I19','J19','K19','L19','M19','N19','O19','P19','Q19','R19','S19','T19','U19',
            'A20','B20','C20','D20','E20','F20','G20','H20','I20','J20','K20','L20','M20','N20','O20','P20','Q20','R20','S20','T20','U20',
            'A21','B21','C21','D21','E21','F21','G21','H21','I21','J21','K21','L21','M21','N21','O21','P21','Q21','R21','S21','T21','U21']

cells_b = ['K4',
           'J5','K5','L5',
           'H6','I6','J6','K6','L6','M6','N6',
           'G7','H7','I7','J7','K7','L7','M7','N7','O7',
           'F8','G8','H8','I8','J8','K8','L8','M8','N8','O8','P8',
           'E9','F9','G9','H9','I9','J9','K9','L9','M9','N9','O9','P9','Q9',
           'E10','G10','O10','Q10',
           'D11','E11','F11','P11','Q11','R11',
           'D12','E12','Q12','R12',
           'C13','D13','R13','S13',
           'C14','D14','R14','S14',
           'C15','S15',
           'C16','S16',
           'C17','S17']

cells_e = ['F10','P10']

cells_w = ['J9','K9','L9',
           'H10','I10','J10','K10','L10','M10','N10',
           'G11','H11','I11','K11','M11','N11','O11',
           'F12','G12','H12','N12','O12','P12',
           'E13','F13','G13','O13','P13','Q13',
           'E14','F14','P14','Q14',
           'D15','E15','F15','H15','J15','L15','N15','P15','Q15','R15',
           'D16','E16','H16','I16','J16','K16','L16','M16','N16','Q16','R16',
           'D17','E17','F17','G17','H17','I17','J17','K17','L17','M17','N17','O17','P17','Q17','R17']

cells_r = ['J11','L11',
           'I12','J12','K12','L12','M12',
           'H13','I13','J13','K13','L13','M13','N13',
           'G14','H14','I14','J14','K14','L14','M14','N14','O14',
           'G15','I15','K15','M15','O15',
           'F16','G16','O16','P16']


for chr in string.ascii_uppercase[:21]:
    sheet.column_dimensions[chr].width = 5
    for i in range(1,22):
        sheet.row_dimensions[1].height = 20
        coord = chr + str(i)
        if coord in cells_lb:
            sheet[coord].fill = fill
        elif coord in cells_b:
            sheet[coord].fill = fill_b
        elif coord in cells_e:
            sheet[coord].fill = fill_e 
        elif coord in cells_w:
            sheet[coord].fill = fill_w
        elif coord in cells_r:
            sheet[coord].fill = fill_r


wb.save('sharkie.xlsx')